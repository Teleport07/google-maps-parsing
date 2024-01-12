import time
import numpy as np
import openpyxl as openpyxl
import requests as requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import openpyxl


options = webdriver.FirefoxOptions()
driver = webdriver.Firefox(options=options)
wait = WebDriverWait(driver, 500)

#подготавливаем excel файл с адресами, по которым нужно получить координаты
wb = openpyxl.load_workbook('путь к excel файлу')
ws = wb['Лист1']
count = 0
driver.get('https://www.google.com/maps/')


for row in range(2, ws.max_row+1):
    if ws.cell(row = row, column = 1).value is not None:
        try:
            #в поисковую строку google maps вносим адрес
            search_bar = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="searchboxinput"]')))
            search_bar.send_keys(ws.cell(row = row, column = 8).value)

            #нажимаем на кнопку поиска
            search_button = driver.find_element(By.XPATH, '//*[@id="searchbox-searchbutton"]')

            search_button.click()
            WebDriverWait(driver, 100).until(
                EC.presence_of_element_located((By.TAG_NAME, "canvas")))

        #если по каким-то причинам не попали на нужную страницу, возвращаемся на предыдущую страницу
        #и снова пишем адрес в поисковую строку
        except:
            driver.execute_script("window.stop()")
            time.sleep(0.5)
            driver.execute_script("window.history.go(-1)")
            time.sleep(3)
            driver.get('https://www.google.com/maps/place/' + ws.cell(row=row, column=1).value)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "canvas")))

        time.sleep(2)
        result_block = driver.find_elements(By.CLASS_NAME, 'fontTitleLarge IFMGgb')

        #если google maps не определил по адресу точное совпадение, то в файл заносим соотвествующую пометку
        if len(result_block) != 0:
            ws.cell(row=row, column=2).value = 'Много результатов, нет точного совпадения'
            count += 1
            time.sleep(np.random.randint(2, 3))
            if count % 10 == 0:
                wb.save('путь к excel файлу')

            WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[2]/div[2]/button'))).click()

        else:
            try:
                #еще один случай, когда google maps не может определить точное совпадение по адресу
                no_results = driver.find_element(By.CLASS_NAME, 'HlvSq')
                if no_results.text == 'Больше результатов нет.':
                    ws.cell(row=row, column=2).value = 'Больше результатов нет.'
                    count += 1
                    time.sleep(np.random.randint(2, 4))
                    if count % 10 == 0:
                        wb.save('путь к excel файлу')
                    WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[2]/div[2]/button'))).click()

                else:
                    #еще один случай не нахождения точного совпадения
                    ws.cell(row=row, column=2).value = 'редкий случай'
            except:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="gb_70"]')))
                ws.cell(row=row, column=2).value = str(
                    driver.find_element(By.XPATH, '//*[@id="gb_70"]').get_attribute('href'))
                count += 1
                time.sleep(np.random.randint(2, 4))
                if count % 10 == 0:
                    wb.save('путь к excel файлу')

                #перед внесением нового адреса в поисковую строку удаляем старый
                try:

                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[2]/div[2]/button')))
                    button_x = driver.find_element(By.XPATH,
                                                   '/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[2]/div[2]/button')
                    driver.execute_script("arguments[0].scrollIntoView();", button_x)
                    driver.execute_script("arguments[0].click();", button_x)
                except:
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH,
                                                        '/ html / body / div[2] / div[3] / div[8] / div[3] / div[1] / div[1] / div / div[2] / div[2] / button')))
                    button_x = driver.find_element(By.XPATH,
                                                   '/ html / body / div[2] / div[3] / div[8] / div[3] / div[1] / div[1] / div / div[2] / div[2] / button')
                    driver.execute_script("arguments[0].scrollIntoView();", button_x)
                    driver.execute_script("arguments[0].click();", button_x)

wb.save('путь к excel файлу')
